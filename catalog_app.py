
import streamlit as st
import pandas as pd
from supabase import create_client
from dotenv import load_dotenv
import os
import io
import base64
import time
import urllib.parse
from datetime import datetime

try:
    from st_keyup import st_keyup
    HAS_KEYUP = True
except ImportError:
    HAS_KEYUP = False



load_dotenv()

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
APP_PASSWORD = os.getenv('APP_PASSWORD', 'TejasImpex@2026')
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_SECONDS = 60

st.set_page_config(
    page_title="TEJAS IMPEX — Product Catalog",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("❌ Supabase credentials not found. Check your .env file.")
    st.stop()

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


if 'authenticated' not in st.session_state:
    st.session_state.authenticated = True
if 'login_attempts' not in st.session_state:
    st.session_state.login_attempts = 0
if 'lockout_until' not in st.session_state:
    st.session_state.lockout_until = 0

def show_login():
    """Show branded login page with security limits."""
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800;900&display=swap');
    *:not([class*="material"]):not([class*="icon"]):not([class*="Icon"]):not([class*="symbol"]):not([class*="Symbol"]):not([data-testid*="Glyph"]):not([data-testid*="Icon"]):not([data-testid*="icon"]) {
        font-family: 'Inter', sans-serif !important;
    }
    .stApp { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); }
    #MainMenu, header, footer { visibility: hidden; }
    .login-container {
        max-width: 420px; margin: 80px auto; padding: 40px;
        background: rgba(255,255,255,0.06); backdrop-filter: blur(20px);
        border-radius: 24px; border: 1px solid rgba(255,255,255,0.1);
        box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    }
    .login-brand { text-align: center; margin-bottom: 32px; }
    .login-logo {
        font-size: 36px; font-weight: 900; color: #e94560;
        letter-spacing: 2px; margin-bottom: 4px;
    }
    .login-sub {
        font-size: 12px; color: rgba(255,255,255,0.35);
        letter-spacing: 3px; text-transform: uppercase;
    }
    .login-title {
        font-size: 22px; font-weight: 700; color: #fff;
        text-align: center; margin-bottom: 8px;
    }
    .login-desc {
        font-size: 13px; color: rgba(255,255,255,0.4);
        text-align: center; margin-bottom: 28px;
    }
    div[data-testid="stTextInput"] input {
        background: rgba(255,255,255,0.08) !important;
        border: 2px solid rgba(255,255,255,0.3) !important;
        color: #fff !important; border-radius: 12px !important;
        padding: 14px 18px !important; font-size: 18px !important;
        font-weight: 600 !important;
    }
    div[data-testid="stTextInput"] input::placeholder { color: rgba(255,255,255,0.65) !important; }
    div[data-testid="stTextInput"] input:focus {
        border-color: #e94560 !important;
        box-shadow: 0 0 20px rgba(233,69,96,0.4) !important;
    }
    div[data-testid="stTextInput"] button {
        color: rgba(255,255,255,0.85) !important;
    }
    div[data-testid="stTextInput"] button:hover {
        color: #ff6b6b !important;
    }
    .stButton > button {
        width: 100%; background: linear-gradient(135deg, #e94560, #ff6b6b) !important;
        color: white !important; border: none !important; border-radius: 12px !important;
        padding: 14px !important; font-size: 16px !important; font-weight: 700 !important;
        letter-spacing: 1px !important; transition: all 0.3s ease !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(233,69,96,0.4) !important;
    }
    .login-footer {
        text-align: center; margin-top: 24px;
        font-size: 11px; color: rgba(255,255,255,0.2);
    }
    .login-error {
        background: rgba(233,69,96,0.15); border: 1px solid rgba(233,69,96,0.3);
        border-radius: 10px; padding: 10px 16px; margin-top: 12px;
        color: #ff6b6b; font-size: 13px; text-align: center;
    }
    .login-locked {
        background: rgba(255,165,0,0.12); border: 1px solid rgba(255,165,0,0.25);
        border-radius: 10px; padding: 10px 16px; margin-top: 12px;
        color: #ffa500; font-size: 13px; text-align: center;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="login-container">
        <div class="login-brand">
            <div class="login-logo">TEJAS IMPEX</div>
            <div class="login-sub">Product Catalog System</div>
        </div>
        <div class="login-title">Welcome Back</div>
        <div class="login-desc">Enter your team password to access the catalog</div>
    </div>
    """, unsafe_allow_html=True)

    # Check lockout
    now = time.time()
    if st.session_state.lockout_until > now:
        remaining = int(st.session_state.lockout_until - now)
        st.markdown(f"""
        <div style="max-width:420px;margin:0 auto;">
            <div class="login-locked">
                Account locked. Try again in {remaining}s
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    col_l, col_c, col_r = st.columns([1, 1.2, 1])
    with col_c:
        pwd = st.text_input("Password", type="password",
                           placeholder="Enter team password",
                           label_visibility="collapsed")
        if st.button("ACCESS CATALOG", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state.authenticated = True
                st.session_state.login_attempts = 0
                st.rerun()
            else:
                st.session_state.login_attempts += 1
                remaining_attempts = MAX_LOGIN_ATTEMPTS - st.session_state.login_attempts
                if remaining_attempts <= 0:
                    st.session_state.lockout_until = time.time() + LOCKOUT_SECONDS
                    st.session_state.login_attempts = 0
                    st.markdown("""
                    <div style="max-width:420px;margin:0 auto;">
                        <div class="login-locked">Too many attempts. Locked for 60 seconds.</div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="max-width:420px;margin:0 auto;">
                        <div class="login-error">Incorrect password. {remaining_attempts} attempts remaining.</div>
                    </div>
                    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="login-footer">
        © 2026 TEJAS IMPEX PVT. LTD. — Authorized Access Only
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# Gate: show login if not authenticated
# if not st.session_state.authenticated:
#     show_login()

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

/* ── Global ── */
*:not([class*="material"]):not([class*="icon"]):not([class*="Icon"]):not([class*="symbol"]):not([class*="Symbol"]):not([data-testid*="Glyph"]):not([data-testid*="Icon"]):not([data-testid*="icon"]) {
    font-family: 'Inter', sans-serif !important;
}

.stApp {
    background: linear-gradient(135deg, #f0f2f5 0%, #e8ecf1 50%, #f5f0f0 100%);
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1a2e 0%, #16213e 40%, #0f3460 100%);
    border-right: none;
}

[data-testid="stSidebar"] * {
    color: #e0e0e0 !important;
}

[data-testid="stSidebar"] input {
    color: #ffffff !important;
    background: rgba(255,255,255,0.1) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
}

[data-testid="stSidebar"] input::placeholder {
    color: rgba(255,255,255,0.45) !important;
}

[data-testid="stSidebar"] label {
    color: #a0aec0 !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 1px !important;
}

[data-testid="stSidebar"] .stSelectbox > div > div {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 10px !important;
    color: #fff !important;
}

[data-testid="stSidebar"] .stSlider > div > div > div {
    background: #e94560 !important;
}

[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.08) !important;
}

/* ── Main Page Inputs & Dropdowns styling ── */
div[data-testid="stSelectbox"] label, div[data-testid="stSlider"] label {
    font-size: 11px !important;
    font-weight: 700 !important;
    color: #4a5568 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.5px !important;
    margin-bottom: 6px !important;
}

div[data-testid="stSelectbox"] > div > div {
    background: white !important;
    border: 2px solid #e8ecf1 !important;
    border-radius: 14px !important;
    color: #1a1a2e !important;
    min-height: 44px !important;
    box-shadow: 0 2px 10px rgba(0,0,0,0.04) !important;
    transition: all 0.3s ease !important;
}

div[data-testid="stSelectbox"] > div > div:hover {
    border-color: #e94560 !important;
}

div[data-testid="stSelectbox"] * {
    color: #1a1a2e !important;
}

div[data-testid="stTextInput"] input {
    background: white !important;
    border: 2px solid #cbd5e1 !important;
    border-radius: 12px !important;
    color: #0f172a !important;
    min-height: 36px !important;
    font-size: 15px !important;
    font-weight: 500 !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04), inset 0 1px 2px rgba(0,0,0,0.05) !important;
    transition: all 0.3s ease !important;
}
div[data-testid="stTextInput"] input::placeholder {
    color: #64748b !important;
    font-size: 14px !important;
}

div[data-testid="stTextInput"] input:focus {
    border-color: #e94560 !important;
    box-shadow: 0 0 0 3px rgba(233, 69, 96, 0.25) !important;
}

/* Reduce top app padding */
[data-testid="stAppViewContainer"] > section > div:first-child {
    padding-top: 0.4rem !important;
}
div.block-container {
    padding-top: 0.6rem !important;
    padding-bottom: 1rem !important;
}

/* ── Header Banner ── */
.hero-banner {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border-radius: 16px;
    padding: 14px 24px;
    margin-bottom: 10px;
    position: relative;
    overflow: hidden;
    box-shadow: 0 6px 24px rgba(26, 26, 46, 0.22);
}

.hero-banner::before {
    content: '';
    position: absolute;
    top: -60%;
    right: -10%;
    width: 300px;
    height: 300px;
    background: radial-gradient(circle, rgba(233,69,96,0.15) 0%, transparent 70%);
    border-radius: 50%;
}

.hero-banner::after {
    content: '';
    position: absolute;
    bottom: -50%;
    left: 20%;
    width: 200px;
    height: 200px;
    background: radial-gradient(circle, rgba(15,52,96,0.3) 0%, transparent 70%);
    border-radius: 50%;
}

.hero-content {
    position: relative;
    z-index: 2;
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.hero-left { display: flex; align-items: center; gap: 14px; }

.hero-logo {
    background: linear-gradient(135deg, #e94560, #ff6b6b);
    color: white;
    font-size: 20px;
    font-weight: 900;
    padding: 8px 14px;
    border-radius: 12px;
    letter-spacing: 1px;
    box-shadow: 0 4px 14px rgba(233,69,96,0.35);
}

.hero-title {
    font-size: 20px;
    font-weight: 800;
    color: #ffffff;
    letter-spacing: 0.5px;
    line-height: 1.2;
}

.hero-sub {
    font-size: 11px;
    color: rgba(255,255,255,0.55);
    font-weight: 400;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-top: 2px;
}

.hero-stats {
    display: flex;
    gap: 20px;
    text-align: center;
}

.hero-stat-value {
    font-size: 22px;
    font-weight: 800;
    color: #ffffff;
    line-height: 1;
}

.hero-stat-label {
    font-size: 10px;
    color: rgba(255,255,255,0.45);
    text-transform: uppercase;
    letter-spacing: 1.2px;
    font-weight: 600;
    margin-top: 4px;
}

.hero-stat-divider {
    width: 1px;
    background: rgba(255,255,255,0.12);
    align-self: stretch;
}

/* ── Search Results Bar ── */
.results-bar {
    background: white;
    border-radius: 10px;
    padding: 8px 18px;
    margin-bottom: 10px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    border: 1px solid rgba(0,0,0,0.04);
}

.results-text {
    font-size: 13px;
    color: #6c757d;
    font-weight: 500;
}

.results-text strong {
    color: #1a1a2e;
    font-weight: 700;
}

.results-accent {
    color: #e94560;
    font-weight: 700;
}

.sort-info {
    font-size: 12px;
    color: #a0aec0;
}

/* ── Animations ── */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(30px); }
    to { opacity: 1; transform: translateY(0); }
}

@keyframes slideInPage {
    from { opacity: 0; transform: translateX(20px); }
    to { opacity: 1; transform: translateX(0); }
}

@keyframes pulseGlow {
    0% { box-shadow: 0 4px 20px rgba(0,0,0,0.06); }
    50% { box-shadow: 0 8px 35px rgba(233,69,96,0.12); }
    100% { box-shadow: 0 4px 20px rgba(0,0,0,0.06); }
}

@keyframes shimmer {
    0% { background-position: -200% 0; }
    100% { background-position: 200% 0; }
}

@keyframes heroFloat {
    0%, 100% { transform: translateY(0px); }
    50% { transform: translateY(-8px); }
}

.stMain > div { animation: slideInPage 0.5s ease-out; }

/* ── Product Card ── */
.product-card {
    background: rgba(255,255,255,0.9);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border-radius: 18px;
    padding: 0;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    border: 1px solid rgba(255,255,255,0.6);
    overflow: hidden;
    height: 100%;
    animation: fadeInUp 0.6s ease-out both;
}

.product-card:hover {
    transform: translateY(-10px) scale(1.02);
    box-shadow: 0 20px 60px rgba(233,69,96,0.18), 0 8px 24px rgba(0,0,0,0.1);
    border-color: rgba(233,69,96,0.35);
}

.card-image-wrap {
    width: 100%;
    height: 200px;
    background: linear-gradient(135deg, #f8f9fa, #f0f2f5);
    display: flex;
    align-items: center;
    justify-content: center;
    overflow: hidden;
    position: relative;
}

.card-image-wrap img {
    max-width: 90%;
    max-height: 90%;
    object-fit: contain;
    transition: transform 0.5s cubic-bezier(0.4, 0, 0.2, 1), filter 0.4s ease;
}

.product-card:hover .card-image-wrap img {
    transform: scale(1.12) rotate(1deg);
    filter: brightness(1.05);
}

.product-card:hover .card-image-wrap {
    background: linear-gradient(135deg, #fff5f5, #f0f0ff);
}

.card-brand-ribbon {
    position: absolute;
    top: 12px;
    left: 0;
    background: linear-gradient(135deg, #e94560, #ff6b6b);
    color: white;
    font-size: 10px;
    font-weight: 700;
    padding: 4px 14px 4px 10px;
    letter-spacing: 1px;
    text-transform: uppercase;
    border-radius: 0 20px 20px 0;
}

.card-body {
    padding: 16px 18px 18px 18px;
}

.card-ref {
    font-size: 12px;
    font-weight: 600;
    color: #a0aec0;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
}

.card-name {
    font-size: 15px;
    font-weight: 700;
    color: #1a1a2e;
    line-height: 1.4;
    margin-bottom: 8px;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    min-height: 42px;
}

.card-price {
    font-size: 24px;
    font-weight: 800;
    color: #1a1a2e;
    margin-bottom: 10px;
}

.card-price .currency {
    font-size: 14px;
    font-weight: 600;
    color: #6c757d;
    vertical-align: super;
    margin-right: 2px;
}

.card-tags {
    display: flex;
    flex-wrap: wrap;
    gap: 5px;
    margin-bottom: 12px;
}

.chip {
    font-size: 10px;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 50px;
    display: inline-block;
    letter-spacing: 0.3px;
}

.chip-cat {
    background: linear-gradient(135deg, #fef0f2, #fce4ec);
    color: #e94560;
}

.chip-series {
    background: linear-gradient(135deg, #e8ecf1, #dce3ed);
    color: #0f3460;
}

.chip-sub {
    background: linear-gradient(135deg, #e8f5e9, #c8e6c9);
    color: #2e7d32;
}

.card-footer {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding-top: 12px;
    border-top: 1px solid rgba(0,0,0,0.05);
}

.card-packing {
    font-size: 11px;
    color: #a0aec0;
    font-weight: 500;
}

.card-packing strong {
    color: #6c757d;
}

/* ── Detail Page ── */
.detail-hero {
    background: linear-gradient(135deg, #1a1a2e, #16213e, #0f3460);
    border-radius: 20px;
    padding: 36px 44px;
    margin-bottom: 28px;
    box-shadow: 0 10px 40px rgba(26,26,46,0.2);
}

.detail-breadcrumb {
    font-size: 13px;
    color: rgba(255,255,255,0.4);
    margin-bottom: 16px;
    font-weight: 500;
}

.detail-breadcrumb a {
    color: rgba(255,255,255,0.6);
    text-decoration: none;
}

.detail-title {
    font-size: 32px;
    font-weight: 800;
    color: #ffffff;
    margin-bottom: 6px;
}

.detail-ref {
    font-size: 16px;
    color: rgba(255,255,255,0.45);
    font-weight: 500;
}

.detail-image-container {
    background: white;
    border-radius: 18px;
    padding: 30px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 400px;
    border: 1px solid rgba(0,0,0,0.04);
}

.detail-image-container img {
    max-width: 100%;
    max-height: 380px;
    object-fit: contain;
}

.detail-info-card {
    background: white;
    border-radius: 18px;
    padding: 28px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    border: 1px solid rgba(0,0,0,0.04);
    margin-bottom: 18px;
}

.detail-price {
    font-size: 38px;
    font-weight: 900;
    color: #1a1a2e;
    margin: 12px 0 20px 0;
}

.detail-price .cur {
    font-size: 20px;
    color: #6c757d;
    font-weight: 600;
}

.detail-meta-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 12px;
    margin: 20px 0;
}

.detail-meta-box {
    background: linear-gradient(135deg, #f8f9fa, #f0f2f5);
    border-radius: 14px;
    padding: 16px;
    text-align: center;
    border: 1px solid rgba(0,0,0,0.04);
}

.detail-meta-num {
    font-size: 26px;
    font-weight: 800;
    color: #1a1a2e;
}

.detail-meta-lbl {
    font-size: 11px;
    font-weight: 700;
    color: #a0aec0;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-top: 4px;
}

.detail-spec-card {
    background: white;
    border-radius: 18px;
    padding: 28px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    border: 1px solid rgba(0,0,0,0.04);
    animation: fadeInUp 0.6s ease-out 0.2s both;
}

.detail-spec-title {
    font-size: 18px;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 16px;
    display: flex;
    align-items: center;
    gap: 10px;
    padding-bottom: 12px;
    border-bottom: 2px solid #f0f2f5;
}

.detail-spec-content {
    background: linear-gradient(135deg, #f8f9fa, #f5f0f0);
    border-radius: 12px;
    padding: 0;
    font-size: 14px;
    line-height: 2;
    color: #2d3436;
    font-family: 'Inter', sans-serif !important;
    border: 1px solid rgba(0,0,0,0.04);
    overflow: hidden;
}

.spec-row {
    display: flex;
    padding: 12px 20px;
    border-bottom: 1px solid #f0f2f5;
    transition: background 0.2s ease;
}

.spec-row:last-child { border-bottom: none; }
.spec-row:hover { background: rgba(233,69,96,0.03); }

.spec-row:nth-child(even) { background: rgba(0,0,0,0.015); }
.spec-row:nth-child(even):hover { background: rgba(233,69,96,0.04); }

.spec-label {
    font-weight: 700;
    color: #1a1a2e;
    min-width: 180px;
    font-size: 13px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.spec-value {
    color: #4a5568;
    font-size: 14px;
    flex: 1;
}

.spec-plain {
    padding: 20px 24px;
    white-space: pre-wrap;
    font-size: 14px;
    line-height: 2;
    color: #2d3436;
}

.detail-badge-row {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    margin-bottom: 14px;
}

.detail-chip {
    font-size: 12px;
    font-weight: 600;
    padding: 6px 16px;
    border-radius: 50px;
}

/* ── Pagination ── */
.pagination-container {
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 6px;
    margin: 30px 0 10px 0;
}

/* ── Sidebar Stat Cards ── */
.sidebar-stat {
    background: rgba(255,255,255,0.06);
    border-radius: 12px;
    padding: 14px 16px;
    margin-bottom: 10px;
    border: 1px solid rgba(255,255,255,0.06);
}

.sidebar-stat-val {
    font-size: 24px;
    font-weight: 800;
    color: #e94560 !important;
}

.sidebar-stat-lbl {
    font-size: 11px;
    font-weight: 600;
    color: rgba(255,255,255,0.4) !important;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* ── Footer ── */
.app-footer {
    text-align: center;
    padding: 28px 20px 18px 20px;
    border-top: 1px solid rgba(0,0,0,0.06);
    margin-top: 40px;
}

.footer-brand {
    font-size: 16px;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 6px;
}

.footer-info {
    font-size: 13px;
    color: #8892b0;
    line-height: 1.8;
}

.footer-copy {
    font-size: 11px;
    color: #c0c0c0;
    margin-top: 10px;
}

/* ── Streamlit Overrides ── */
.stButton > button {
    border-radius: 12px !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1) !important;
    font-family: 'Inter', sans-serif !important;
}

.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(0,0,0,0.12) !important;
}

div[data-testid="stComponentSandbox"] iframe {
    height: 36px !important;
}

div[data-testid="stTextInput"] input {
    border-radius: 12px !important;
    border: 2px solid #cbd5e1 !important;
    padding: 4px 12px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    color: #0f172a !important;
    transition: all 0.3s ease !important;
    background: white !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04), inset 0 1px 2px rgba(0,0,0,0.05) !important;
}
div[data-testid="stTextInput"] input::placeholder {
    color: #64748b !important;
    font-size: 12px !important;
}

div[data-testid="stTextInput"] input:focus {
    border-color: #e94560 !important;
    box-shadow: 0 0 0 3px rgba(233, 69, 96, 0.25) !important;
}

/* ── Logo in banner ── */
.hero-logo-img {
    height: 40px;
    border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.2);
    transition: transform 0.4s ease;
    background: white;
    padding: 4px;
}

.hero-logo-img:hover {
    transform: scale(1.08);
}

.hero-logos {
    display: flex;
    gap: 10px;
    align-items: center;
}

/* Sleek styling for Expander */
div[data-testid="stExpander"] {
    background: white !important;
    border-radius: 12px !important;
    border: 2px solid #cbd5e1 !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
    overflow: hidden !important;
    margin-bottom: 10px !important;
    transition: all 0.3s ease !important;
}
div[data-testid="stExpander"]:hover {
    border-color: #a0aec0 !important;
    box-shadow: 0 4px 14px rgba(0,0,0,0.06) !important;
}
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] {
    background: #f8f9fa !important;
    padding: 14px 16px !important;
    border-top: 1px solid #e8ecf1 !important;
}
div[data-testid="stExpander"] summary {
    font-weight: 700 !important;
    font-size: 13px !important;
    color: #1a1a2e !important;
    padding: 8px 14px !important;
    transition: all 0.3s ease !important;
    background: #ffffff !important;
}
div[data-testid="stExpander"] summary:hover {
    color: #e94560 !important;
    background: rgba(233,69,96,0.03) !important;
}
div[data-testid="stExpander"] summary svg {
    fill: #e94560 !important;
}

/* Hide Streamlit elements */
#MainMenu { visibility: hidden; }
header { visibility: hidden; }
footer { visibility: hidden; }

/* ── No Image Placeholder ── */
.no-image-placeholder {
    width: 100%;
    height: 100%;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    color: #c0c0c0;
    font-size: 14px;
    gap: 8px;
}

.no-image-icon {
    font-size: 48px;
    opacity: 0.4;
}

/* ── WhatsApp Sharing Styles ── */
.whatsapp-btn {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    width: 100%;
    background: linear-gradient(135deg, #25d366, #128c7e) !important;
    color: white !important;
    font-weight: 700;
    font-size: 14px;
    padding: 12px 20px;
    border-radius: 12px;
    border: none;
    cursor: pointer;
    box-shadow: 0 4px 15px rgba(37, 211, 102, 0.15);
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    text-decoration: none;
    margin-top: 14px;
}
.whatsapp-btn:hover {
    transform: translateY(-2px) !important;
    color: white !important;
    box-shadow: 0 6px 22px rgba(37, 211, 102, 0.3) !important;
}

.card-wa-btn {
    background: #eafbee !important;
    color: #128c7e !important;
    border: 1px solid rgba(37, 211, 102, 0.2) !important;
    box-shadow: none !important;
    padding: 8px !important;
    font-size: 13px !important;
    margin-top: 8px !important;
}
.card-wa-btn:hover {
    background: #25d366 !important;
    color: white !important;
    box-shadow: 0 4px 12px rgba(37, 211, 102, 0.2) !important;
}

/* ── Custom scrollbars ── */
div.main::-webkit-scrollbar {
    width: 10px !important;
    height: 10px !important;
    display: block !important;
}
div.main::-webkit-scrollbar-track {
    background: #f1f2f5 !important;
}
div.main::-webkit-scrollbar-thumb {
    background: #cbd5e1 !important;
    border-radius: 5px !important;
    border: 2px solid #f1f2f5 !important;
}
div.main::-webkit-scrollbar-thumb:hover {
    background: #94a3b8 !important;
}
</style>

<!-- Floating Go to Top Button -->
<div id="scroll-to-top-container">
    <button id="goToTopBtn" title="Go to Top" style="
        position: fixed;
        bottom: 30px;
        right: 30px;
        z-index: 999999;
        width: 48px;
        height: 48px;
        border-radius: 50%;
        background: #e94560;
        color: white;
        border: none;
        box-shadow: 0 4px 15px rgba(233, 69, 96, 0.4);
        cursor: pointer;
        display: none;
        align-items: center;
        justify-content: center;
        transition: all 0.3s ease;
        opacity: 0.8;
    ">
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round">
            <polyline points="18 15 12 9 6 15"></polyline>
        </svg>
    </button>
</div>

<script>
    (function() {
        const doc = window.parent.document;
        const initScrollBtn = () => {
            const mainEl = doc.querySelector('.main');
            if (!mainEl) {
                setTimeout(initScrollBtn, 100);
                return;
            }
            
            let btn = doc.getElementById('goToTopBtn');
            if (!btn) {
                const container = doc.createElement('div');
                container.id = 'scroll-to-top-container';
                
                const newBtn = doc.createElement('button');
                newBtn.id = 'goToTopBtn';
                newBtn.innerHTML = '<svg xmlns=\"http://www.w3.org/2000/svg\" width=\"24\" height=\"24\" viewBox=\"0 0 24 24\" fill=\"none\" stroke=\"currentColor\" stroke-width=\"3\" stroke-linecap=\"round\" stroke-linejoin=\"round\"><polyline points=\"18 15 12 9 6 15\"></polyline></svg>';
                newBtn.style.cssText = 'position:fixed; bottom:30px; right:30px; z-index:999999; width:48px; height:48px; border-radius:50%; background:#e94560; color:white; border:none; box-shadow:0 4px 15px rgba(233, 69, 96, 0.4); cursor:pointer; display:none; align-items:center; justify-content:center; transition:all 0.3s cubic-bezier(0.4, 0, 0.2, 1); opacity:0.8; outline:none;';
                
                newBtn.onmouseover = function() {
                    newBtn.style.transform = 'translateY(-4px) scale(1.05)';
                    newBtn.style.opacity = '1';
                    newBtn.style.boxShadow = '0 6px 20px rgba(233, 69, 96, 0.6)';
                    newBtn.style.background = '#ff6b6b';
                };
                newBtn.onmouseout = function() {
                    newBtn.style.transform = 'none';
                    newBtn.style.opacity = '0.8';
                    newBtn.style.boxShadow = '0 4px 15px rgba(233, 69, 96, 0.4)';
                    newBtn.style.background = '#e94560';
                };
                newBtn.onclick = function() {
                    mainEl.scrollTo({top: 0, behavior: 'smooth'});
                };
                container.appendChild(newBtn);
                doc.body.appendChild(container);
                btn = newBtn;
            }
            
            mainEl.removeEventListener('scroll', mainEl._scrollSpyHandler);
            mainEl._scrollSpyHandler = function() {
                if (mainEl.scrollTop > 300) {
                    btn.style.display = 'flex';
                } else {
                    btn.style.display = 'none';
                }
            };
            mainEl.addEventListener('scroll', mainEl._scrollSpyHandler);
        };
        
        if (doc.readyState === 'complete' || doc.readyState === 'interactive') {
            initScrollBtn();
        } else {
            doc.addEventListener('DOMContentLoaded', initScrollBtn);
        }
    })();
</script>
""", unsafe_allow_html=True)


defaults = {
    'view': 'catalog',       # 'catalog' or 'detail'
    'selected_product': None,
    'page': 1,
    'search_term': '',
    'category': 'All',
    'sub_category': 'All',
    'company': 'All',
    'series': 'All',
    'sort_by': 'Name A→Z',
    'price_range': (0, 100000),
    'selected_products': set(),  # set of product ids selected for bulk share
}

for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val


@st.cache_data(ttl=300)
def load_products():
    try:
        all_data = []
        offset = 0
        batch_size = 1000
        while True:
            response = supabase.table('products').select('*').range(offset, offset + batch_size - 1).execute()
            if response.data:
                all_data.extend(response.data)
                if len(response.data) < batch_size:
                    break
                offset += batch_size
            else:
                break
        if all_data:
            df = pd.DataFrame(all_data)
            # Clean numeric columns
            df['mrp'] = pd.to_numeric(df['mrp'], errors='coerce').fillna(0)
            df['packing_pcs'] = pd.to_numeric(df['packing_pcs'], errors='coerce').fillna(0).astype(int)
            df['packing_bx'] = pd.to_numeric(df['packing_bx'], errors='coerce').fillna(0).astype(int)
            df['packing_car'] = pd.to_numeric(df['packing_car'], errors='coerce').fillna(0).astype(int)
            # Clean text columns
            for col in ['category', 'sub_category', 'company', 'series', 'product_name', 'ref_code', 'specification']:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str).str.strip()
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading products: {e}")
        return pd.DataFrame()

with st.spinner('Loading catalog...'):
    df = load_products()

if df.empty:
    st.warning("⚠️ No products found. Check your Supabase connection.")
    st.stop()


def make_toggle(pid, key):
    def toggle_cb():
        if st.session_state.get(key):
            st.session_state.selected_products.add(pid)
        else:
            st.session_state.selected_products.discard(pid)
    return toggle_cb


def format_price(price):
    """Format price with commas (Indian style)."""
    if price == 0:
        return "—"
    price_str = f"{price:,.2f}"
    return price_str

def get_unique_sorted(series):
    """Get sorted unique non-empty values."""
    return sorted([v for v in series.dropna().unique().tolist() if v and str(v).strip()])

def load_logo_base64(filename):
    """Load a logo image as base64 for embedding in HTML."""
    filepath = os.path.join(os.path.dirname(__file__), filename)
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    return None

# Load logos
tejas_logo_b64 = load_logo_base64('Tejas.png')
deli_logo_b64 = load_logo_base64('delitools-logo (1).png')

def format_spec_html(spec_text):
    """Parse specification text into styled HTML rows."""
    if not spec_text or not str(spec_text).strip():
        return ''
    lines = str(spec_text).strip().split('\n')
    rows_html = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Try to parse key: value or key - value patterns
        for sep in [':', ' - ', '–', '\t']:
            if sep in line:
                parts = line.split(sep, 1)
                if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                    rows_html.append(
                        f'<div class="spec-row">'
                        f'<div class="spec-label">{parts[0].strip()}</div>'
                        f'<div class="spec-value">{parts[1].strip()}</div>'
                        f'</div>'
                    )
                    break
        else:
            # No separator found, render as plain row
            rows_html.append(
                f'<div class="spec-row">'
                f'<div class="spec-value" style="min-width:100%">{line}</div>'
                f'</div>'
            )
    return '\n'.join(rows_html) if rows_html else f'<div class="spec-plain">{spec_text}</div>'

def get_whatsapp_share_url(product):
    """Generate a clean, professional WhatsApp share URL."""
    name = product.get('product_name', 'Unknown')
    spec_text = product.get('specification', '')
    ref = product.get('ref_code', '-')
    mrp = product.get('mrp', 0)

    message = f"*Product:* {name}\n"
    
    if spec_text and str(spec_text).strip():
        message += f"*Specs:* {str(spec_text).strip()}\n"
        
    message += (
        f"*Model Number:* {ref}\n"
        f"*MRP:* Rs. {mrp:,.2f}\n"
    )

    encoded_message = urllib.parse.quote(message.strip())
    return f"https://wa.me/?text={encoded_message}"

def generate_pdf_report(products_list):
    """Generate a structured PDF for selected products using fpdf2.
    Draw order per card: background fill → accent bar → image → text (so nothing covers text).
    """

    def safe_text(s):
        """Replace characters outside Latin-1 so Helvetica does not crash."""
        if not s:
            return ''
        return ''.join(c if ord(c) < 256 else '?' for c in str(s))

    try:
        from fpdf import FPDF
        import urllib.request
        import tempfile, os

        # ── Layout constants ──────────────────────────────────────
        MARGIN      = 10      # left/right page margin (mm)
        PAGE_W      = 190     # usable width  (210 - 2*10)
        IMG_X       = MARGIN + 4          # image left edge
        IMG_W       = 48      # image column width (mm)
        TEXT_X      = IMG_X + IMG_W + 5   # text column start
        TEXT_W      = PAGE_W - IMG_W - 12 # text column width
        CARD_H      = 58      # fixed card height (mm)
        IMG_H       = CARD_H - 6          # image height inside card
        CARD_GAP    = 5       # vertical gap between cards

        class PDF(FPDF):
            def header(self):
                # Dark navy header bar
                self.set_fill_color(26, 26, 46)
                self.rect(0, 0, 210, 20, 'F')
                # Company name
                self.set_font('Helvetica', 'B', 14)
                self.set_text_color(255, 255, 255)
                self.set_xy(0, 4)
                self.cell(150, 8, 'TEJAS IMPEX PVT. LTD.', align='L', border=0)
                # Subtitle
                self.set_font('Helvetica', '', 8)
                self.set_text_color(180, 190, 220)
                self.set_xy(0, 13)
                self.cell(150, 5, 'Product Catalog Report', align='L', border=0)
                # Date on right
                from datetime import date
                self.set_font('Helvetica', '', 8)
                self.set_text_color(200, 210, 230)
                self.set_xy(140, 8)
                self.cell(60, 6, date.today().strftime('%d %B %Y'), align='R', border=0)
                self.set_text_color(0, 0, 0)
                self.ln(5)

            def footer(self):
                self.set_y(-12)
                self.set_font('Helvetica', 'I', 8)
                self.set_text_color(150, 150, 150)
                self.cell(0, 8,
                    f'Page {self.page_no()} | tejasimpex2023@gmail.com | +977-9801986465 | Teku, Kathmandu',
                    align='C', border=0)

        pdf = PDF()
        pdf.set_auto_page_break(auto=False)   # we handle page breaks manually
        pdf.set_margins(MARGIN, 24, MARGIN)
        pdf.add_page()

        TOP_Y   = 26   # first card starts below header
        BOTTOM_Y = 282  # page bottom limit (A4 = 297mm minus footer)

        cur_y = TOP_Y

        for idx, p in enumerate(products_list):
            import textwrap
            # ── Collect + sanitise data ──────────────────────────
            name     = safe_text(p.get('product_name', 'Unknown'))
            ref      = safe_text(p.get('ref_code', '-'))
            brand    = safe_text(p.get('company', ''))
            series   = safe_text(p.get('series', ''))
            category = safe_text(p.get('category', ''))
            sub_cat  = safe_text(p.get('sub_category', ''))
            spec     = safe_text(p.get('specification', ''))
            mrp      = p.get('mrp', 0)
            pcs      = p.get('packing_pcs', 0)
            bx       = p.get('packing_bx', 0)
            image_url = p.get('image_url', '')

            # Parse spec into multiple wrapped lines
            spec_lines = []
            if spec and str(spec).strip():
                for line in str(spec).strip().split('\n'):
                    line = line.strip()
                    if line:
                        wrapped = textwrap.wrap(line, width=70)
                        spec_lines.extend(wrapped)

            # Calculate dynamic card height
            n_spec_lines = len(spec_lines)
            spec_section_h = n_spec_lines * 4.5
            
            # Dynamic calculation of text height before specification and packing
            name_lines = 2 if len(name) > 38 else 1
            non_spec_h = 5 + (name_lines * 6) + 3 + 6 + 6  # margins + name + divider + ref + category
            if pcs or bx:
                non_spec_h += 6  # include packing spacing
                
            card_h = max(55, non_spec_h + spec_section_h + 17) # 17mm buffer for MRP badge & margin
            img_h = card_h - 6

            # ── Page-break check ────────────────────────────────
            if cur_y + card_h > BOTTOM_Y:
                pdf.add_page()
                cur_y = TOP_Y

            # ── 1. Draw card background (FIRST) ─────────────────
            pdf.set_fill_color(248, 249, 252)
            pdf.set_draw_color(220, 226, 238)
            pdf.set_line_width(0.3)
            pdf.rect(MARGIN, cur_y, PAGE_W, card_h, style='FD')

            # ── 2. Red accent bar on left ────────────────────────
            pdf.set_fill_color(233, 69, 96)
            pdf.rect(MARGIN, cur_y, 3, card_h, style='F')

            # ── 3. Serial number chip (top-right corner) ─────────
            pdf.set_fill_color(26, 26, 46)
            pdf.rect(MARGIN + PAGE_W - 14, cur_y + 2, 12, 7, style='F')
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(MARGIN + PAGE_W - 14, cur_y + 2)
            pdf.cell(12, 7, f'#{idx+1:02d}', align='C', border=0)

            # ── 4. Fetch & draw image (SECOND) ───────────────────
            img_placed = False
            if image_url and str(image_url).strip():
                try:
                    url_lower = image_url.lower()
                    suffix = '.png' if '.png' in url_lower else \
                             '.webp' if '.webp' in url_lower else '.jpg'
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
                    urllib.request.urlretrieve(image_url, tmp.name)
                    tmp.close()
                    # White image bg
                    pdf.set_fill_color(255, 255, 255)
                    pdf.rect(IMG_X, cur_y + 3, IMG_W, img_h, style='F')
                    pdf.image(tmp.name, x=IMG_X + 1, y=cur_y + 4, w=IMG_W - 2, h=img_h - 2)
                    os.unlink(tmp.name)
                    img_placed = True
                except Exception:
                    pass

            if not img_placed:
                # Grey placeholder
                pdf.set_fill_color(230, 232, 238)
                pdf.rect(IMG_X, cur_y + 3, IMG_W, img_h, style='F')
                pdf.set_font('Helvetica', '', 8)
                pdf.set_text_color(160, 165, 180)
                pdf.set_xy(IMG_X, cur_y + 3 + img_h/2 - 3)
                pdf.cell(IMG_W, 6, 'No Image', align='C', border=0)

            # ── 5. Write text (THIRD – always visible over bg) ───
            ty = cur_y + 5  # text top

            # Product name (bold, wraps up to 2 lines)
            pdf.set_font('Helvetica', 'B', 11)
            pdf.set_text_color(26, 26, 46)
            pdf.set_xy(TEXT_X, ty)
            # Truncate name to avoid overflow
            max_name = name[:72] + ('...' if len(name) > 72 else '')
            pdf.multi_cell(TEXT_W - 14, 6, max_name, border=0, align='L')
            ty = pdf.get_y() + 1

            # ── Divider line ──────────────────────────────────────
            pdf.set_draw_color(200, 210, 230)
            pdf.set_line_width(0.2)
            pdf.line(TEXT_X, ty, TEXT_X + TEXT_W - 14, ty)
            ty += 3

            # Row: Ref + Brand
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(90, 100, 130)
            pdf.set_xy(TEXT_X, ty)
            pdf.cell(22, 5, 'Ref Code:', border=0)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(40, 50, 80)
            pdf.cell(TEXT_W/2 - 22, 5, ref, border=0)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(90, 100, 130)
            pdf.cell(18, 5, 'Brand:', border=0)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(40, 50, 80)
            pdf.cell(TEXT_W/2 - 18, 5, brand, border=0)
            ty += 6

            # Row: Category + Series
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(90, 100, 130)
            pdf.set_xy(TEXT_X, ty)
            pdf.cell(22, 5, 'Category:', border=0)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(40, 50, 80)
            cat_str = category + (f' > {sub_cat}' if sub_cat else '')
            pdf.cell(TEXT_W/2 - 22, 5, cat_str[:38], border=0)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(90, 100, 130)
            pdf.cell(18, 5, 'Series:', border=0)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(40, 50, 80)
            pdf.cell(TEXT_W/2 - 18, 5, series[:30], border=0)
            ty += 6

            # Row: Specifications (multiple lines)
            if spec_lines:
                pdf.set_font('Helvetica', 'B', 8)
                pdf.set_text_color(90, 100, 130)
                pdf.set_xy(TEXT_X, ty)
                pdf.cell(22, 4, 'Spec:', border=0)
                
                pdf.set_font('Helvetica', '', 8)
                pdf.set_text_color(60, 70, 100)
                for sline in spec_lines:
                    pdf.set_xy(TEXT_X + 22, ty)
                    pdf.cell(TEXT_W - 22, 4, sline, border=0)
                    ty += 4.5
                ty += 1.5

            # Row: Packing
            packing_str = ''
            if pcs:
                packing_str += f'{pcs} pcs/unit'
            if bx:
                packing_str += f'   {bx} units/box'
            if packing_str:
                pdf.set_font('Helvetica', 'B', 8)
                pdf.set_text_color(90, 100, 130)
                pdf.set_xy(TEXT_X, ty)
                pdf.cell(22, 5, 'Packing:', border=0)
                pdf.set_font('Helvetica', '', 8)
                pdf.set_text_color(40, 50, 80)
                pdf.cell(TEXT_W - 22, 5, packing_str, border=0)
                ty += 6

            # ── MRP badge at bottom of card ───────────────────────
            mrp_y = cur_y + card_h - 13
            # Red pill background
            pdf.set_fill_color(233, 69, 96)
            pdf.rect(TEXT_X, mrp_y, 55, 9, style='F', round_corners=True, corner_radius=2.0)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(TEXT_X + 2, mrp_y + 1)
            pdf.cell(51, 7, f'MRP: Rs. {mrp:,.2f}', align='C', border=0)

            # Tax note
            pdf.set_font('Helvetica', 'I', 7)
            pdf.set_text_color(120, 130, 160)
            pdf.set_xy(TEXT_X + 58, mrp_y + 2)
            pdf.cell(50, 5, 'Incl. all taxes', border=0)

            # ── Advance Y ────────────────────────────────────────
            cur_y += card_h + CARD_GAP

        # ── Output ───────────────────────────────────────────────
        pdf_bytes = pdf.output()
        return bytes(pdf_bytes)

    except ImportError:
        # Fallback HTML
        rows = ""
        for p in products_list:
            name      = p.get('product_name', 'Unknown')
            ref       = p.get('ref_code', '-')
            mrp       = p.get('mrp', 0)
            category  = p.get('category', '')
            sub_cat   = p.get('sub_category', '')
            brand     = p.get('company', '')
            series    = p.get('series', '')
            spec      = p.get('specification', '')
            pcs       = p.get('packing_pcs', 0)
            bx        = p.get('packing_bx', 0)
            image_url = p.get('image_url', '')
            img_html  = f'<img src="{image_url}" style="width:130px;height:auto;object-fit:contain;border-radius:8px;" />' if image_url else ''
            pack      = (f'{pcs} pcs' + (f' | {bx}/box' if bx else '')) if pcs else ''
            rows += f"""
            <tr style="border-bottom:2px solid #e8ecf4;">
              <td style="padding:14px;width:150px;vertical-align:top;">{img_html}</td>
              <td style="padding:14px;vertical-align:top;">
                <div style="font-size:16px;font-weight:700;color:#1a1a2e;margin-bottom:6px;">{name}</div>
                <table style="font-size:12px;color:#555;border-collapse:collapse;">
                  <tr><td style="padding:2px 8px 2px 0;font-weight:600;color:#888;">Ref Code</td><td style="padding:2px;">{ref}</td>
                      <td style="padding:2px 8px 2px 12px;font-weight:600;color:#888;">Brand</td><td style="padding:2px;">{brand}</td></tr>
                  <tr><td style="padding:2px 8px 2px 0;font-weight:600;color:#888;">Category</td><td style="padding:2px;">{category}{(' > ' + sub_cat) if sub_cat else ''}</td>
                      <td style="padding:2px 8px 2px 12px;font-weight:600;color:#888;">Series</td><td style="padding:2px;">{series}</td></tr>
                  {'<tr><td style="padding:2px 8px 2px 0;font-weight:600;color:#888;">Spec</td><td colspan=3 style="padding:2px;">' + spec[:120] + '</td></tr>' if spec else ''}
                  {'<tr><td style="padding:2px 8px 2px 0;font-weight:600;color:#888;">Packing</td><td colspan=3 style="padding:2px;">' + pack + '</td></tr>' if pack else ''}
                </table>
                <div style="margin-top:10px;display:inline-block;background:#e94560;color:white;
                     font-size:15px;font-weight:700;padding:6px 16px;border-radius:6px;">
                  Rs. {mrp:,.2f} <span style="font-size:10px;font-weight:400;opacity:.85">incl. taxes</span>
                </div>
              </td>
            </tr>"""
        html = (
            "<!DOCTYPE html><html><head><meta charset='utf-8'>"
            "<style>body{font-family:Arial,sans-serif;max-width:860px;margin:0 auto;padding:24px;}"
            "h1{background:#1a1a2e;color:white;padding:16px 24px;border-radius:10px;margin-bottom:20px;}"
            "table.main{width:100%;border-collapse:collapse;}</style></head>"
            f"<body><h1>TEJAS IMPEX PVT. LTD. &mdash; Product Catalog</h1>"
            f"<table class='main'>{rows}</table></body></html>"
        )
        return html.encode('utf-8')



def export_to_excel(df_to_export):
    """Format catalog columns and export as Excel binary data (fallback to CSV if openpyxl lacks)."""
    # Select columns
    cols = [
        'ref_code', 'product_name', 'category', 'sub_category',
        'company', 'series', 'mrp', 'packing_pcs', 'packing_bx',
        'packing_car', 'specification', 'image_url'
    ]
    # Filter columns that actually exist
    valid_cols = [c for c in cols if c in df_to_export.columns]
    export_df = df_to_export[valid_cols].copy()
    
    # Rename columns to human-readable names
    rename_dict = {
        'ref_code': 'Model/Reference',
        'product_name': 'Product Name',
        'category': 'Category',
        'sub_category': 'Sub-Category',
        'company': 'Company/Brand',
        'series': 'Series',
        'mrp': 'MRP (NPR)',
        'packing_pcs': 'Packing (Pcs)',
        'packing_bx': 'Packing (Box)',
        'packing_car': 'Packing (Carton)',
        'specification': 'Specifications',
        'image_url': 'Image URL'
    }
    export_df = export_df.rename(columns={k: v for k, v in rename_dict.items() if k in export_df.columns})
    
    # Try openpyxl
    try:
        import openpyxl
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Product Catalog')
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    except Exception:
        # Fallback to CSV
        csv_str = export_df.to_csv(index=False)
        return csv_str.encode('utf-8'), "text/csv", "csv"




if st.session_state.view == 'detail' and st.session_state.selected_product is not None:
    p = st.session_state.selected_product

    if st.button("← Back to Catalog", type="primary"):
        st.session_state.view = 'catalog'
        st.session_state.selected_product = None
        st.rerun()

    # Hero
    st.markdown(f"""
    <div class="detail-hero">
        <div class="detail-breadcrumb">
            Catalog &nbsp;›&nbsp; {p.get('category', '')} &nbsp;›&nbsp; {p.get('sub_category', '')}
        </div>
        <div class="detail-title">{p.get('product_name', 'Product')}</div>
        <div class="detail-ref">Model: {p.get('ref_code', '—')}</div>
    </div>
    """, unsafe_allow_html=True)

    # Two-column layout
    col_img, col_info = st.columns([1, 1], gap="large")

    with col_img:
        image_url = p.get('image_url', '')
        if image_url and str(image_url).strip():
            st.markdown(f"""
            <div class="detail-image-container">
                <img src="{image_url}" alt="{p.get('product_name', '')}">
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="detail-image-container">
                <div class="no-image-placeholder">
                    <div class="no-image-icon">📦</div>
                    <div>No image available</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

    with col_info:
        # Info Card
        company = p.get('company', '')
        series = p.get('series', '')
        category = p.get('category', '')
        sub_cat = p.get('sub_category', '')
        mrp = p.get('mrp', 0)

        tags_html = f'<span class="detail-chip chip-cat">{category}</span>' if category else ''
        if sub_cat:
            tags_html += f' <span class="detail-chip chip-sub">{sub_cat}</span>'
        if series:
            tags_html += f' <span class="detail-chip chip-series">{series}</span>'
        if company:
            tags_html += f' <span class="detail-chip" style="background:linear-gradient(135deg,#fff3e0,#ffe0b2);color:#e65100;">{company}</span>'

        st.markdown(f"""
        <div class="detail-info-card">
            <div class="detail-badge-row">{tags_html}</div>
            <div class="detail-price"><span class="cur">रु</span> {format_price(mrp)}</div>
            <div style="font-size: 13px; color: #2d6a4f; font-weight: 600; margin-bottom: 16px;">
                ✅ MRP (Inclusive of all taxes)
            </div>
            <div class="detail-meta-grid">
                <div class="detail-meta-box">
                    <div class="detail-meta-num">{p.get('packing_pcs', 0)}</div>
                    <div class="detail-meta-lbl">Per Piece</div>
                </div>
                <div class="detail-meta-box">
                    <div class="detail-meta-num">{p.get('packing_bx', 0)}</div>
                    <div class="detail-meta-lbl">Per Box</div>
                </div>
                <div class="detail-meta-box">
                    <div class="detail-meta-num">{p.get('packing_car', 0)}</div>
                    <div class="detail-meta-lbl">Per Carton</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        wa_url = get_whatsapp_share_url(p)
        st.markdown(f"""
        <a href="{wa_url}" target="_blank" style="text-decoration: none;">
            <button class="whatsapp-btn">
                <svg viewBox="0 0 24 24" width="18" height="18" fill="currentColor" style="display:inline-block; vertical-align:-3px; margin-right:4px;">
                    <path d="M12.031 2c-5.514 0-9.99 4.476-9.99 9.99 0 2.057.625 3.993 1.8 5.617L2 22l4.56-1.328a9.92 9.92 0 005.47 1.62c5.514 0 9.99-4.476 9.99-9.99C22.02 6.476 17.545 2 12.031 2zm6.657 14.394c-.255.708-1.5 1.29-2.07 1.35-.55.06-1.27.08-2.03-.16-.62-.2-1.35-.45-2.09-.78-2.95-1.3-4.85-4.32-5-4.52-.15-.2-1.12-1.49-1.12-2.84 0-1.35.7-2.01.95-2.28.25-.26.56-.33.74-.33.18 0 .37 0 .53.01.17 0 .39-.06.6.46.22.54.76 1.85.83 1.99.07.14.12.31.02.51-.1.2-.15.31-.3.48-.15.17-.32.39-.46.52-.16.15-.33.31-.14.63.19.32.85 1.4 1.82 2.27.97.87 1.78 1.13 2.1 1.29.32.16.51.13.7-.09.19-.22.82-.95 1.04-1.28.22-.33.44-.28.75-.17.3.11 1.93.91 2.27 1.08.33.17.56.25.64.39.08.14.08.82-.17 1.53z"/>
                </svg>
                Share details on WhatsApp
            </button>
        </a>
        """, unsafe_allow_html=True)

    # Specifications (parsed into styled rows)
    spec = p.get('specification', '')
    spec_html = format_spec_html(spec)
    if spec_html:
        st.markdown(f"""
        <div class="detail-spec-card">
            <div class="detail-spec-title">📋 Specifications</div>
            <div class="detail-spec-content">{spec_html}</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="detail-spec-card">
            <div class="detail-spec-title">📋 Specifications</div>
            <div class="detail-spec-content">
                <div class="spec-plain" style="color:#a0aec0; text-align:center;">
                    No specifications available for this product.
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Source info (subtle)
    src_file = p.get('source_file', '')
    src_sheet = p.get('source_sheet', '')
    if src_file or src_sheet:
        st.markdown(f"""
        <div style="margin-top: 16px; padding: 12px 18px; background: #f8f9fa; border-radius: 12px;
                    font-size: 12px; color: #a0aec0;">
            📁 Source: {src_file} {f"/ {src_sheet}" if src_sheet else ""}
        </div>
        """, unsafe_allow_html=True)



elif st.session_state.view == 'catalog':
    # ── Stats Calculations ──
    categories_count = len(get_unique_sorted(df['category']))
    companies_count = len(get_unique_sorted(df['company']))

    # Hero Banner with logos
    logos_html = '<div class="hero-logos">'
    if tejas_logo_b64:
        logos_html += f'<img src="data:image/png;base64,{tejas_logo_b64}" class="hero-logo-img" alt="Tejas Impex">'
    if deli_logo_b64:
        logos_html += f'<img src="data:image/png;base64,{deli_logo_b64}" class="hero-logo-img" alt="Deli Tools">'
    logos_html += '</div>'

    st.markdown(f"""
    <div class="hero-banner">
        <div class="hero-content">
            <div class="hero-left">
                {logos_html}
                <div>
                    <div class="hero-title">TEJAS IMPEX PVT. LTD.</div>
                    <div class="hero-sub">Premium Product Catalog</div>
                </div>
            </div>
            <div class="hero-stats">
                <div>
                    <div class="hero-stat-value">{len(df)}</div>
                    <div class="hero-stat-label">Total Products</div>
                </div>
                <div class="hero-stat-divider"></div>
                <div>
                    <div class="hero-stat-value">{categories_count}</div>
                    <div class="hero-stat-label">Categories</div>
                </div>
                <div class="hero-stat-divider"></div>
                <div>
                    <div class="hero-stat-value">{companies_count}</div>
                    <div class="hero-stat-label">Brands</div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Top Search Input (Full Width Layout) ──
    if HAS_KEYUP:
        search = st_keyup(
            "Search",
            value=st.session_state.search_term,
            placeholder="🔍 Search products by name, reference code, specs...",
            label_visibility="collapsed",
            debounce=250,
            key="keyup_search_input"
        )
    else:
        search = st.text_input(
            "Search",
            value=st.session_state.search_term,
            placeholder="🔍 Search products by name, reference code, specs...",
            label_visibility="collapsed",
            key="text_search_input"
        )
    if search != st.session_state.search_term:
        st.session_state.search_term = search
        st.session_state.page = 1

    # Callback to safely reset all filters before widgets are instantiated in the subsequent run
    def reset_filters_callback():
        max_p = float(df['mrp'].max()) if df['mrp'].max() > 0 else 100000.0
        st.session_state.search_term = ''
        st.session_state.category = 'All'
        st.session_state.sub_category = 'All'
        st.session_state.company = 'All'
        st.session_state.series = 'All'
        st.session_state.sort_by = 'Name A→Z'
        st.session_state.price_range = (0.0, max_p)
        st.session_state.page = 1
        
        # Safely overwrite widget session state keys prior to script instantiation
        st.session_state.keyup_search_input = ''
        st.session_state.text_search_input = ''
        st.session_state.cat_selectbox = 'All Catalog'
        st.session_state.sub_cat_selectbox = 'All Sub-Categories'
        st.session_state.company_selectbox = 'All Brands'
        st.session_state.series_selectbox = 'All Series'
        st.session_state.price_slider = (0.0, max_p)
        st.session_state.sort_selectbox = 'Name A→Z'

    # Collapsible Filter and Sort Options
    with st.expander("🎛️ Filter & Sort Options", expanded=False):
        # Row 1: Categories, Sub-Categories, Companies, Series selectors
        col_cat, col_sub, col_comp, col_series = st.columns(4, gap="small")
        
        with col_cat:
            categories = get_unique_sorted(df['category'])
            cat = st.selectbox(
                "📂 Category", ["All Catalog"] + categories,
                index=(["All Catalog"] + categories).index("All Catalog" if st.session_state.category == "All" else st.session_state.category) if st.session_state.category in (["All"] + categories) else 0,
                key="cat_selectbox"
            )
            cat_val = "All" if cat == "All Catalog" else cat
            if cat_val != st.session_state.category:
                st.session_state.category = cat_val
                st.session_state.sub_category = 'All'
                st.session_state.page = 1

        with col_sub:
            if st.session_state.category != 'All':
                sub_cats_df = df[df['category'] == st.session_state.category]
            else:
                sub_cats_df = df
            sub_categories = get_unique_sorted(sub_cats_df['sub_category'])
            sub_cat = st.selectbox(
                "📁 Sub-Category", ["All Sub-Categories"] + sub_categories,
                index=(["All Sub-Categories"] + sub_categories).index("All Sub-Categories" if st.session_state.sub_category == "All" else st.session_state.sub_category) if st.session_state.sub_category in (["All"] + sub_categories) else 0,
                key="sub_cat_selectbox"
            )
            sub_cat_val = "All" if sub_cat == "All Sub-Categories" else sub_cat
            if sub_cat_val != st.session_state.sub_category:
                st.session_state.sub_category = sub_cat_val
                st.session_state.page = 1

        with col_comp:
            companies = get_unique_sorted(df['company'])
            comp = st.selectbox(
                "🏢 Brand", ["All Brands"] + companies,
                index=(["All Brands"] + companies).index("All Brands" if st.session_state.company == "All" else st.session_state.company) if st.session_state.company in (["All"] + companies) else 0,
                key="company_selectbox"
            )
            comp_val = "All" if comp == "All Brands" else comp
            if comp_val != st.session_state.company:
                st.session_state.company = comp_val
                st.session_state.page = 1

        with col_series:
            series_list = get_unique_sorted(df['series'])
            series_options = ["All Series"] + series_list if series_list else ["All Series"]
            ser = st.selectbox(
                "🎨 Series", series_options,
                index=series_options.index("All Series" if st.session_state.series == "All" else st.session_state.series) if st.session_state.series in (["All"] + series_list) else 0,
                key="series_selectbox"
            )
            ser_val = "All" if ser == "All Series" else ser
            if ser_val != st.session_state.series:
                st.session_state.series = ser_val
                st.session_state.page = 1

        # Row 2: Price Range Slider & Sort By Selection
        col_price, col_sort = st.columns([6.0, 4.0], gap="medium")
        
        with col_price:
            max_price = float(df['mrp'].max()) if df['mrp'].max() > 0 else 100000.0
            curr_low, curr_high = st.session_state.price_range
            curr_low = float(max(0.0, min(float(curr_low), max_price)))
            curr_high = float(max(0.0, min(float(curr_high), max_price)))
            if curr_low > curr_high:
                curr_low, curr_high = 0.0, max_price
                
            price_range = st.slider(
                "💰 Price Range (रु)",
                min_value=0.0,
                max_value=max_price,
                value=(curr_low, curr_high),
                format="रु%.0f",
                key="price_slider"
            )
            if price_range != st.session_state.price_range:
                st.session_state.price_range = (float(price_range[0]), float(price_range[1]))
                st.session_state.page = 1

        with col_sort:
            sort_options = ["Name A→Z", "Name Z→A", "Price ↑ Low to High", "Price ↓ High to Low", "Ref Code"]
            sort = st.selectbox(
                "↕️ Sort By", sort_options,
                index=sort_options.index(st.session_state.sort_by) if st.session_state.sort_by in sort_options else 0,
                key="sort_selectbox"
            )
            st.session_state.sort_by = sort

        # Row 3: Action Buttons (Reset and Export)
        col_reset, col_export = st.columns(2, gap="medium")
        with col_reset:
            st.button("🔄 Reset Filters", use_container_width=True, key="reset_filters_btn", on_click=reset_filters_callback)

        # Apply filtering logic inside the expander context to export correct list
        filtered = df.copy()
        
        if st.session_state.search_term:
            q = st.session_state.search_term.lower()
            filtered = filtered[
                filtered['product_name'].str.lower().str.contains(q, na=False) |
                filtered['ref_code'].str.lower().str.contains(q, na=False) |
                filtered['category'].str.lower().str.contains(q, na=False) |
                filtered['sub_category'].str.lower().str.contains(q, na=False) |
                filtered['specification'].str.lower().str.contains(q, na=False)
            ]
            
        if st.session_state.category != 'All':
            filtered = filtered[filtered['category'] == st.session_state.category]
        if st.session_state.sub_category != 'All':
            filtered = filtered[filtered['sub_category'] == st.session_state.sub_category]
        if st.session_state.company != 'All':
            filtered = filtered[filtered['company'] == st.session_state.company]
        if st.session_state.series != 'All':
            filtered = filtered[filtered['series'] == st.session_state.series]
            
        p_low, p_high = st.session_state.price_range
        filtered = filtered[(filtered['mrp'] >= p_low) & (filtered['mrp'] <= p_high)]
        
        # Sort
        sort_col = st.session_state.sort_by
        if sort_col == "Name A→Z":
            filtered = filtered.sort_values('product_name', ascending=True)
        elif sort_col == "Name Z→A":
            filtered = filtered.sort_values('product_name', ascending=False)
        elif sort_col == "Price ↑ Low to High":
            filtered = filtered.sort_values('mrp', ascending=True)
        elif sort_col == "Price ↓ High to Low":
            filtered = filtered.sort_values('mrp', ascending=False)
        elif sort_col == "Ref Code":
            filtered = filtered.sort_values('ref_code', ascending=True)

        with col_export:
            excel_data, mime_type, file_ext = export_to_excel(filtered)
            st.download_button(
                label="📥 Export List",
                data=excel_data,
                file_name=f"TEJAS_IMPEX_Catalog_{datetime.now().strftime('%Y%m%d')}.{file_ext}",
                mime=mime_type,
                use_container_width=True,
                key="export_btn_catalog"
            )

    # Results bar
    total_results = len(filtered)
    active_filters = []
    if st.session_state.category != 'All':
        active_filters.append(st.session_state.category)
    if st.session_state.sub_category != 'All':
        active_filters.append(st.session_state.sub_category)
    if st.session_state.company != 'All':
        active_filters.append(st.session_state.company)
    if st.session_state.series != 'All':
        active_filters.append(st.session_state.series)
    if st.session_state.search_term:
        active_filters.append(f'"{st.session_state.search_term}"')

    filter_text = " in " + ", ".join(active_filters) if active_filters else ""
    n_selected = len(st.session_state.selected_products)

    # Pre-build selected badge to avoid broken f-string HTML
    selected_badge = ''
    if n_selected:
        selected_badge = f'  &bull;  <span style="color:#e94560;font-weight:700;">{n_selected} selected</span>'

    st.markdown(f"""
    <div class="results-bar">
        <div class="results-text">
            Showing <span class="results-accent">{total_results}</span> products{filter_text}{selected_badge}
        </div>
        <div class="sort-info">Sorted by: {st.session_state.sort_by}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Bulk Share Controls (shown when items are selected) ──
    if n_selected > 0:
        sel_col1, sel_col2, sel_col3 = st.columns([2.5, 2.5, 1.5])
        with sel_col1:
            st.markdown(
                f'<div style="padding:6px 0;font-size:13px;color:#1a1a2e;font-weight:600;">'
                f'📋 {n_selected} product(s) selected</div>',
                unsafe_allow_html=True
            )
        with sel_col2:
            # Build list of selected product dicts
            sel_products = []
            for pid in st.session_state.selected_products:
                matches = df[df['id'] == pid]
                if not matches.empty:
                    sel_products.append(matches.iloc[0].to_dict())

            sel_ids_key = ",".join(sorted(str(pid) for pid in st.session_state.selected_products))
            pdf_ready_key = st.session_state.get('pdf_ready_key', '')

            if pdf_ready_key != sel_ids_key:
                if st.button(f"⚙️ Prepare PDF ({n_selected} items)", key="prep_pdf_btn", use_container_width=True):
                    with st.spinner("Preparing PDF layout..."):
                        try:
                            pdf_bytes = generate_pdf_report(sel_products)
                            st.session_state.pdf_ready_bytes = pdf_bytes
                            st.session_state.pdf_ready_ext = "pdf"
                            st.session_state.pdf_ready_mime = "application/pdf"
                        except Exception as e:
                            # Fallback HTML on any error
                            rows_fb = ""
                            for p in sel_products:
                                nm = p.get('product_name', '')
                                rf = p.get('ref_code', '')
                                pr = p.get('mrp', 0)
                                img = p.get('image_url', '')
                                img_tag = f'<img src="{img}" style="width:100px;height:auto;"/>' if img else ''
                                rows_fb += f'<tr><td style="padding:8px;">{img_tag}</td><td style="padding:8px;"><b>{nm}</b><br>Ref: {rf}<br>Rs. {pr:,.2f}</td></tr>'
                            html_fb = f'<html><body style="font-family:sans-serif"><h2>TEJAS IMPEX PVT. LTD.</h2><table border="0" cellspacing="8">{rows_fb}</table></body></html>'
                            st.session_state.pdf_ready_bytes = html_fb.encode('utf-8')
                            st.session_state.pdf_ready_ext = "html"
                            st.session_state.pdf_ready_mime = "text/html"
                        st.session_state.pdf_ready_key = sel_ids_key
                        st.rerun()
            else:
                st.download_button(
                    label=f"📥 Download PDF ({n_selected} items)",
                    data=st.session_state.pdf_ready_bytes,
                    file_name=f"TEJAS_IMPEX_Selected_{datetime.now().strftime('%Y%m%d_%H%M')}.{st.session_state.pdf_ready_ext}",
                    mime=st.session_state.pdf_ready_mime,
                    use_container_width=True,
                    key="bulk_pdf_download"
                )
        with sel_col3:
            if st.button("✖ Clear Selection", key="clear_sel_btn", use_container_width=True):
                st.session_state.selected_products = set()
                if 'pdf_ready_key' in st.session_state:
                    del st.session_state.pdf_ready_key
                st.rerun()

    # ── Pagination Setup removed: listing all items ──
    page_data = filtered
    total_products = len(filtered)
    start_idx = 0
    end_idx = total_products

    # ── Product Grid ──
    if len(page_data) == 0:
        st.markdown("""
        <div style="text-align: center; padding: 80px 20px;">
            <div style="font-size: 64px; margin-bottom: 16px;">🔍</div>
            <div style="font-size: 22px; font-weight: 700; color: #1a1a2e;">No products found</div>
            <div style="font-size: 15px; color: #6c757d; margin-top: 8px;">
                Try adjusting your search or filters to find what you're looking for
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        # Render in rows of 4
        rows = [page_data.iloc[i:i+4] for i in range(0, len(page_data), 4)]

        for row_chunk in rows:
            cols = st.columns(4, gap="medium")
            for idx, (_, product) in enumerate(row_chunk.iterrows()):
                with cols[idx]:
                    image_url = product.get('image_url', '')
                    if not image_url or not str(image_url).strip():
                        img_html = """
                        <div class="no-image-placeholder">
                            <div class="no-image-icon">📦</div>
                        </div>"""
                    else:
                        img_html = f'<img src="{image_url}" alt="{product.get("product_name", "")}" loading="lazy">'

                    company = product.get('company', '')
                    ref_code = product.get('ref_code', '')
                    name = product.get('product_name', 'Unknown')
                    mrp = product.get('mrp', 0)
                    category = product.get('category', '')
                    series = product.get('series', '')
                    sub_cat = product.get('sub_category', '')
                    pcs = product.get('packing_pcs', 0)
                    bx = product.get('packing_bx', 0)

                    tags = ""
                    if category:
                        tags += f'<span class="chip chip-cat">{category}</span>'
                    if series:
                        tags += f'<span class="chip chip-series">{series}</span>'
                    if sub_cat:
                        tags += f'<span class="chip chip-sub">{sub_cat}</span>'

                    packing_info = ""
                    if pcs:
                        packing_info += f"<strong>{pcs}</strong> pcs"
                    if bx:
                        packing_info += f" · <strong>{bx}</strong> /box"

                    st.markdown(f"""
                    <div class="product-card">
                        <div class="card-image-wrap">
                            {f'<div class="card-brand-ribbon">{company}</div>' if company else ''}
                            {img_html}
                        </div>
                        <div class="card-body">
                            <div class="card-ref">{ref_code}</div>
                            <div class="card-name">{name}</div>
                            <div class="card-price">
                                <span class="currency">रु</span>{format_price(mrp)}
                            </div>
                            <div class="card-tags">{tags}</div>
                            <div class="card-footer">
                                <div class="card-packing">{packing_info if packing_info else '—'}</div>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    # 3-column button row: Details | Share | Select checkbox
                    prod_id = product.get('id')
                    is_selected = prod_id in st.session_state.selected_products
                    col_btn1, col_btn2, col_btn3 = st.columns([1.2, 1.0, 0.6])

                    with col_btn1:
                        if st.button("\U0001f50e Details", key=f"view_{prod_id}", use_container_width=True):
                            st.session_state.selected_product = product.to_dict()
                            st.session_state.view = 'detail'
                            st.rerun()

                    with col_btn2:
                        wa_url = get_whatsapp_share_url(product)
                        st.markdown(f"""
                        <a href="{wa_url}" target="_blank" style="text-decoration:none;display:block;">
                            <button style="
                                width:100%;height:38px;display:flex;align-items:center;
                                justify-content:center;gap:4px;border-radius:10px;
                                background:#25D366;color:white;border:none;
                                font-family:'Inter',sans-serif;font-size:13px;
                                font-weight:600;cursor:pointer;
                                transition:all 0.2s ease;box-sizing:border-box;
                            ">\U0001f4f2 Share</button>
                        </a>
                        """, unsafe_allow_html=True)

                    with col_btn3:
                        st.checkbox(
                            "Select",
                            value=is_selected,
                            key=f"chk_{prod_id}",
                            label_visibility="collapsed",
                            on_change=make_toggle(prod_id, f"chk_{prod_id}")
                        )



    # ── Footer ──
    st.markdown("""
    <div class="app-footer">
        <div class="footer-brand">TEJAS IMPEX PVT. LTD.</div>
        <div class="footer-info">
            📍 Teku, Kathmandu, Nepal &nbsp;·&nbsp;
            ✉️ tejasimpex2023@gmail.com &nbsp;·&nbsp;
            📱 9801986465
        </div>
        <div class="footer-copy">© 2026 TEJAS IMPEX PVT. LTD. All Rights Reserved</div>
    </div>
    """, unsafe_allow_html=True)

